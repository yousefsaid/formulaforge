from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.main import app


@pytest.fixture
def client() -> TestClient:
    with TestClient(app) as test_client:
        yield test_client


def workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 2
    sheet["A2"] = 3
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_apply_formula_writes_validated_formula_into_workbook(client: TestClient) -> None:
    response = client.post(
        "/api/v1/workbooks/apply-formula",
        data={"sheet_name": "Sheet", "target_cell": "B2", "formula": "=SUM(A1:A2)"},
        files={
            "file": (
                "book.xlsx",
                workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    assert response.headers["content-disposition"].startswith("attachment")
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["Sheet"]["B2"].value == "=SUM(A1:A2)"


def test_apply_formula_rejects_formula_that_fails_validation(client: TestClient) -> None:
    response = client.post(
        "/api/v1/workbooks/apply-formula",
        data={"sheet_name": "Sheet", "target_cell": "B2", "formula": "=SUM(A1:A9)"},
        files={
            "file": (
                "book.xlsx",
                workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 400
    assert "no longer validates" in response.json()["detail"]


def test_apply_formula_rejects_non_xlsx(client: TestClient) -> None:
    response = client.post(
        "/api/v1/workbooks/apply-formula",
        data={"sheet_name": "Sheet", "target_cell": "B2", "formula": "=SUM(A1:A2)"},
        files={"file": ("book.txt", b"not a workbook", "text/plain")},
    )
    assert response.status_code == 415
