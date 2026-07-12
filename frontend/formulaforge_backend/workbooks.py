from __future__ import annotations

import io
import zipfile

from fastapi import HTTPException
from openpyxl import load_workbook

from .contracts import CellValue, WorkbookPreview

MAX_BYTES = 2 * 1024 * 1024
MAX_SHEETS = 10
MAX_CELLS = 5000
MAX_EXPANDED_BYTES = 20 * 1024 * 1024


def normalize_workbook(raw: bytes, selected_sheet: str | None = None) -> WorkbookPreview:
    if len(raw) > MAX_BYTES:
        raise HTTPException(413, "Workbook exceeds the 2 MB limit")
    try:
        archive = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile as exc:
        raise HTTPException(400, "Invalid .xlsx archive") from exc
    names = archive.namelist()
    if any(name.endswith("vbaProject.bin") for name in names):
        raise HTTPException(400, "Macro-enabled workbooks are not supported")
    if any("externalLinks/" in name for name in names):
        raise HTTPException(400, "External links are not supported")
    if sum(item.file_size for item in archive.infolist()) > MAX_EXPANDED_BYTES:
        raise HTTPException(413, "Workbook archive expands beyond the safety limit")
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=False, keep_links=False)
    except Exception as exc:
        raise HTTPException(
            400, "Workbook could not be opened; encrypted files are not supported"
        ) from exc
    if len(workbook.sheetnames) > MAX_SHEETS:
        raise HTTPException(400, "Workbook has too many sheets")
    sheet_name = selected_sheet or workbook.sheetnames[0]
    if sheet_name not in workbook.sheetnames:
        raise HTTPException(400, "Selected sheet does not exist")
    worksheet = workbook[sheet_name]
    cells: list[CellValue] = []
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cells.append(CellValue(address=cell.coordinate.upper(), value=cell.value))
                if len(cells) > MAX_CELLS:
                    raise HTTPException(400, "Workbook context has too many non-empty cells")
    return WorkbookPreview(
        sheets=workbook.sheetnames,
        selected_sheet=sheet_name,
        cells=cells,
        non_empty_cell_count=len(cells),
    )
